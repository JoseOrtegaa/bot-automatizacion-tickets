from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime

NAME_WORKBOOK = './registers/test_tasker.xlsx'

# Iniciamos conexion con WebDriver Selenium
def setup(URL_TEST):
    #Iniciamos el WebDriver
    driver = webdriver.Chrome()
    #Llamamos a la pagine web
    driver.get(URL_TEST)
    return driver

# Salimos de la conexion
def teardown(driver):
    driver.quit()

## Logica para leer la excel y selenium.
def executeReadExcel(ws):
        try:
            print(f'{datetime.now()} - [TASKER - Init executeReadExcel]')
            driver = ''
            for col in ws.iter_cols(min_row=1, max_col=1):
                for i, cells in enumerate(col, start=1):
                    # Verificar si la celda tiene un hipervínculo
                    if cells.hyperlink:
                        date = datetime.now()
                        url_ticket = cells.hyperlink.target
                        driver = setup(url_ticket)
                        status = excelModify(driver)
                        driver.close()
                        cells_selected_status = ws[f'B{i}']
                        cells_selected_status.value = status

                        cells_selected_update_date = ws[f'C{i}']
                        updated = datetime.now()
                        cells_selected_update_date.value = updated
                        print(f'Indice:{i} ,Status:{status}, Date :{updated}')

        except Exception as e:
            print(f'{datetime.now()} - [TASKER - ERROR AL REALIZAR LA CONSULTA {e}]')
        finally:
            print(f'{datetime.now()} - [TASKER - Fin executeReadExcel]')
            teardown(driver)

        
                    
def excelModify(driver):
    #Obtenemos elemento SELECT
    current_status = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "current_status_ticket")))
    selector_elements = Select(current_status)
    #Obtenemos la opcion selecionada
    selected_option = selector_elements.first_selected_option
    return selected_option.text
             
dateInitTasker = datetime.now()            
print(f'{datetime.now()} - [TASKER - Init]')
# Iniciamos el libro de trabajo de openpyxl
def init_worbook(NAME_WORKBOOK):
    return load_workbook(filename = NAME_WORKBOOK)

## Init ##
wb = init_worbook(NAME_WORKBOOK)
ws = wb.active

# Metodo que ejecuta la logcia de leer excel, abrir selenium y editar la excel.
executeReadExcel(ws)
print(f'{datetime.now()} - [TASKER - Fin]')

## Guardamos los cambios
wb.save(NAME_WORKBOOK)







